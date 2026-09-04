"""Safe, bounded parsing for client spreadsheets exported by other systems."""
import csv
import io
import re
import unicodedata
import zipfile
from datetime import date, datetime

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


MAX_FILE_BYTES = 2 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 12 * 1024 * 1024
MAX_ROWS = 200
MAX_COLUMNS = 50
ALIASES = {
    "name": {"nome", "nomecompleto", "cliente", "razaosocial", "razao"},
    "email": {"email", "correioeletronico", "e-mail"},
    "phone": {"telefone", "celular", "whatsapp", "fone"},
    "tax_id": {"cpf", "cnpj", "cpfcnpj", "documento", "documentofiscal"},
    "stage": {"etapa", "situacao", "status", "fase"},
}


def _key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(r"[^a-z0-9]", "", "".join(char for char in text if not unicodedata.combining(char)).lower())


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()[:1000]


def _table(rows: list[list[object]]) -> dict:
    if not rows or not rows[0]:
        raise HTTPException(422, "A planilha não possui cabeçalho.")
    headers = [_cell(value)[:120] for value in rows[0]]
    if len(headers) > MAX_COLUMNS or any(not header for header in headers) or len({_key(header) for header in headers}) != len(headers):
        raise HTTPException(422, "Use até 50 colunas com nomes únicos e preenchidos na primeira linha.")
    data = []
    for values in rows[1:MAX_ROWS + 1]:
        record = {header: _cell(values[index] if index < len(values) else None) for index, header in enumerate(headers)}
        if any(record.values()):
            data.append(record)
    if len(rows) > MAX_ROWS + 1:
        raise HTTPException(422, "Importe no máximo 200 clientes por arquivo.")
    if not data:
        raise HTTPException(422, "A planilha não possui clientes para importar.")
    suggested = {}
    for field, aliases in ALIASES.items():
        match = next((header for header in headers if _key(header) in {_key(alias) for alias in aliases}), None)
        if match:
            suggested[field] = match
    return {"columns": headers, "rows": data, "row_count": len(data), "suggested_mapping": suggested}


def _csv(content: bytes) -> list[list[object]]:
    text = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None or "\x00" in text:
        raise HTTPException(422, "O CSV deve estar em UTF-8 ou no formato padrão do Excel para Windows.")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row[:MAX_COLUMNS + 1] for row in csv.reader(io.StringIO(text), dialect)][:MAX_ROWS + 2]


def _xlsx(content: bytes) -> list[list[object]]:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > 1000 or sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_BYTES:
                raise HTTPException(422, "A planilha compactada excede o limite seguro.")
            if any(entry.compress_size and entry.file_size / entry.compress_size > 200 for entry in entries):
                raise HTTPException(422, "A planilha possui compactação incompatível com a importação segura.")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True, max_row=MAX_ROWS + 2):
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        workbook.close()
        return rows
    except HTTPException:
        raise
    except (OSError, ValueError, KeyError, zipfile.BadZipFile, InvalidFileException):
        raise HTTPException(422, "O arquivo XLSX não pôde ser lido. Exporte novamente sem senha ou macros.") from None


def parse_client_file(filename: str, content: bytes) -> dict:
    if not content or len(content) > MAX_FILE_BYTES:
        raise HTTPException(413, "Use um arquivo CSV ou XLSX de até 2 MB.")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        rows = _csv(content)
    elif suffix == "xlsx":
        rows = _xlsx(content)
    else:
        raise HTTPException(422, "Formato não aceito. Use CSV ou XLSX.")
    if any(len(row) > MAX_COLUMNS for row in rows):
        raise HTTPException(422, "A planilha excede 50 colunas.")
    return _table(rows)
